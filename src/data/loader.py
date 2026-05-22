import pandas as pd
import numpy as np
from typing import Union, Optional, Dict, Any, Tuple
import logging
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

logger = logging.getLogger(__name__)

class DataLoader:
    """Handles loading of data from a single Excel workbook with multiple sheets."""

    @staticmethod
    def load_workbook(file_path: str) -> Dict[str, pd.DataFrame]:
        """
        Load all sheets from the Excel workbook.

        Args:
            file_path: Path to the Excel workbook

        Returns:
            Dictionary with sheet names as keys and DataFrames as values
        """
        try:
            xl = pd.ExcelFile(file_path)
            sheets_data = {}

            for sheet_name in xl.sheet_names:
                try:
                    # Load each sheet with appropriate header handling
                    sheet_lower = sheet_name.lower().strip()
                    
                    # Check for Drop Dump sheet variations
                    if any(variant in sheet_lower for variant in ['drop dump', 'dropdump', 'drop_dump', 'drop-dump']) or sheet_lower == 'drop':
                        # Drop dump has headers on row 0
                        df = pd.read_excel(xl, sheet_name=sheet_name, header=0)
                    elif 'scheme' in sheet_lower:
                        # Scheme sheet: try to detect header row automatically
                        temp_df = pd.read_excel(xl, sheet_name=sheet_name, header=None, nrows=10)
                        
                        header_row = 0
                        for idx, row in temp_df.iterrows():
                            row_str = ' '.join([str(val).lower() for val in row if pd.notna(val)])
                            if 'master model' in row_str or 'master' in row_str:
                                header_row = idx
                                break
                        
                        df = pd.read_excel(xl, sheet_name=sheet_name, header=header_row)
                        logger.info(f"Scheme sheet loaded with header at row {header_row}")
                    elif 'sales' in sheet_lower:
                        # Sales sheet: detect header row automatically
                        temp_df = pd.read_excel(xl, sheet_name=sheet_name, header=None, nrows=10)
                        
                        header_row = 0
                        for idx, row in temp_df.iterrows():
                            row_str = ' '.join([str(val).lower() for val in row if pd.notna(val)])
                            if 'imei' in row_str or 'master model' in row_str or 'sell out' in row_str:
                                header_row = idx
                                logger.info(f"Sales sheet header detected at row {header_row}")
                                break
                        
                        df = pd.read_excel(xl, sheet_name=sheet_name, header=header_row)
                        logger.info(f"Sales sheet loaded with header at row {header_row}, shape: {df.shape}")
                    elif 'price' in sheet_lower:
                        # Price list: detect header row automatically
                        temp_df = pd.read_excel(xl, sheet_name=sheet_name, header=None, nrows=10)
                        
                        header_row = 0
                        for idx, row in temp_df.iterrows():
                            row_str = ' '.join([str(val).lower() for val in row if pd.notna(val)])
                            if 'master model' in row_str or 'model' in row_str:
                                header_row = idx
                                break
                        
                        df = pd.read_excel(xl, sheet_name=sheet_name, header=header_row)
                        logger.info(f"Price list loaded with header at row {header_row}")
                    else:
                        # Other sheets
                        df = pd.read_excel(xl, sheet_name=sheet_name, header=None)

                    sheets_data[sheet_name.lower()] = df
                    logger.info(f"Loaded sheet '{sheet_name}' with shape {df.shape}")

                except Exception as e:
                    logger.warning(f"Error loading sheet '{sheet_name}': {str(e)}")
                    sheets_data[sheet_name.lower()] = pd.DataFrame()

            return sheets_data

        except Exception as e:
            logger.error(f"Error loading workbook {file_path}: {str(e)}")
            raise

    @staticmethod
    def extract_data_sources(workbook_data: Dict[str, pd.DataFrame]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """
        Extract the four required data sources from the workbook.

        Args:
            workbook_data: Dictionary of all sheets from the workbook

        Returns:
            Tuple of (sales_data, price_list, scheme_file, drop_dump)
        """
        logger.info(f"Available sheets in workbook: {list(workbook_data.keys())}")

        # Extract sales data - try multiple possible sheet names
        sales_data = None
        for sheet_name in workbook_data.keys():
            if 'sales' in sheet_name.lower():
                sales_data = workbook_data[sheet_name]
                logger.info(f"Found sales data in sheet: {sheet_name}")
                break

        if sales_data is None or sales_data.empty:
            logger.error(f"Sales sheet not found. Available sheets: {list(workbook_data.keys())}")
            raise ValueError(f"Sales sheet not found in workbook. Available sheets: {list(workbook_data.keys())}")

        # # Extract price list - COMMENTED OUT - NOT USED
        # # Client provides all prices in Sales sheet
        # price_list = None
        # for sheet_name in workbook_data.keys():
        #     if 'price' in sheet_name.lower():
        #         price_list = workbook_data[sheet_name]
        #         logger.info(f"Found price list in sheet: {sheet_name} (will be ignored)")
        #         break
        # 
        # if price_list is None or price_list.empty:
        #     logger.warning(f"Price List sheet not found or empty. Available sheets: {list(workbook_data.keys())}")
        #     logger.warning("Processing will continue without Price List validation")
        #     price_list = pd.DataFrame()  # Create empty dataframe
        
        # Create empty price list dataframe - not used in processing
        price_list = pd.DataFrame()
        logger.info("Price List sheet skipped - all prices come from Sales sheet")

        # Extract scheme file - try multiple possible sheet names
        scheme_file = None
        for sheet_name in workbook_data.keys():
            if 'scheme' in sheet_name.lower():
                scheme_file = workbook_data[sheet_name]
                logger.info(f"Found scheme data in sheet: {sheet_name}")
                break

        if scheme_file is None or scheme_file.empty:
            logger.error(f"Scheme sheet not found. Available sheets: {list(workbook_data.keys())}")
            raise ValueError(f"Scheme sheet not found in workbook. Available sheets: {list(workbook_data.keys())}")

        # Extract drop dump - try multiple possible sheet names
        drop_dump = None
        for sheet_name in workbook_data.keys():
            sheet_lower = sheet_name.lower().strip()
            # Check for various drop dump sheet name variations
            if any(variant in sheet_lower for variant in ['drop dump', 'dropdump', 'drop_dump', 'drop-dump']) or sheet_lower == 'drop':
                drop_dump = workbook_data[sheet_name]
                logger.info(f"Found drop dump in sheet: {sheet_name}")
                break

        if drop_dump is None or drop_dump.empty:
            logger.error(f"Drop Dump sheet not found. Available sheets: {list(workbook_data.keys())}")
            raise ValueError(f"Drop Dump sheet not found in workbook. Available sheets: {list(workbook_data.keys())}")

        logger.info(f"Data shapes - Sales: {sales_data.shape}, Price: {price_list.shape}, Scheme: {scheme_file.shape}, Drop: {drop_dump.shape}")

        # Standardize column names
        sales_data = DataLoader._standardize_sales_columns(sales_data)
        # price_list = DataLoader._standardize_price_columns(price_list)  # COMMENTED OUT - NOT USED
        scheme_file = DataLoader._standardize_scheme_columns(scheme_file)
        drop_dump = DataLoader._standardize_drop_columns(drop_dump)

        return sales_data, price_list, scheme_file, drop_dump

    @staticmethod
    def _standardize_sales_columns(df: pd.DataFrame) -> pd.DataFrame:
        """Standardize column names for sales data."""
        # Convert all column names to strings first, then lowercase for consistent mapping
        df.columns = df.columns.astype(str).str.lower().str.strip()
        
        logger.info(f"Sales sheet original columns: {list(df.columns)}")

        column_mapping = {
            'imei': 'IMEI',
            'sell out date': 'Sell_Out_Date',
            'sellout date': 'Sell_Out_Date',
            'sell_out_date': 'Sell_Out_Date',
            'master modal': 'Master_Model',
            'master_modal': 'Master_Model',
            'master model': 'Master_Model',
            'mastermodel': 'Master_Model',
            'master_model': 'Master_Model',
            'distibutor': 'Distributor',
            'distributor': 'Distributor',
            'purchase date': 'Purchase_Date',
            'purchasedate': 'Purchase_Date',
            'purchase_date': 'Purchase_Date',
            'purchase price': 'Purchase_Price',
            'purchaseprice': 'Purchase_Price',
            'purchase_price': 'Purchase_Price',
            'pur cost mop': 'Purchase_Price',  # Variation for Purchase Price / MOP
            'bill less in invoice': 'Current_Month_Invoice_Price',
            'bill less in invoice ': 'Current_Month_Invoice_Price',
            'current month invoice price': 'Current_Month_Invoice_Price',
            'invoice price (pre gst amount': 'Current_Month_Pre_GST_Invoice_Price',
            'invoice price (pre gst amount)': 'Current_Month_Pre_GST_Invoice_Price',
            'current month pre-gst of invoice price': 'Current_Month_Pre_GST_Invoice_Price',
            'series': 'SERIES',
            'drop': 'Original_Drop',  # Rename existing drop column to avoid conflicts
            'current mop/srp': 'Current_MOP_SRP',  # Standardized name
            'activation date': 'Activation_Date',  # Standardized name
            'final price for multiplication': 'Final_Price_For_Multiplication',
            'final price for multipliaction': 'Final_Price_For_Multiplication',  # Handle typo
            'final_price_for_multiplication': 'Final_Price_For_Multiplication',
            'finalprice for multiplication': 'Final_Price_For_Multiplication',
            'final price multiplication': 'Final_Price_For_Multiplication',
        }

        df = df.rename(columns=column_mapping)
        
        logger.info(f"Sales sheet columns after mapping: {list(df.columns)}")

        # Ensure required columns exist - if not found, try fuzzy matching
        required_cols = ['IMEI', 'Sell_Out_Date', 'Master_Model', 'Distributor']
        missing_cols = []
        
        for col in required_cols:
            if col not in df.columns:
                missing_cols.append(col)
                # Try to find similar columns by partial matching
                if col == 'IMEI':
                    for c in df.columns:
                        if 'imei' in str(c).lower():
                            df = df.rename(columns={c: 'IMEI'})
                            logger.info(f"Mapped '{c}' to 'IMEI'")
                            missing_cols.remove(col)
                            break
                elif col == 'Sell_Out_Date':
                    for c in df.columns:
                        if 'sell' in str(c).lower() and 'date' in str(c).lower():
                            df = df.rename(columns={c: 'Sell_Out_Date'})
                            logger.info(f"Mapped '{c}' to 'Sell_Out_Date'")
                            missing_cols.remove(col)
                            break
                elif col == 'Master_Model':
                    for c in df.columns:
                        if ('master' in str(c).lower() or 'model' in str(c).lower()) and 'mop' not in str(c).lower():
                            df = df.rename(columns={c: 'Master_Model'})
                            logger.info(f"Mapped '{c}' to 'Master_Model'")
                            missing_cols.remove(col)
                            break
                elif col == 'Distributor':
                    for c in df.columns:
                        if 'distrib' in str(c).lower():
                            df = df.rename(columns={c: 'Distributor'})
                            logger.info(f"Mapped '{c}' to 'Distributor'")
                            missing_cols.remove(col)
                            break

        if missing_cols:
            logger.error(f"Required columns still missing after fuzzy matching: {missing_cols}")
            logger.error(f"Available columns: {list(df.columns)}")
            raise ValueError(f"Missing required columns in sales data: {missing_cols}. Available columns: {list(df.columns)[:10]}")
        
        # Round all price/amount columns in Sales sheet to whole numbers (except percentages)
        price_amount_columns = [
            'Purchase_Price', 'Current_Month_Invoice_Price', 'Current_Month_Pre_GST_Invoice_Price',
            'Current_MOP_SRP', 'Original_Drop', 'Purchase_Price'  # MOP at the Time of Purchase maps to Purchase_Price
        ]
        for col in price_amount_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').round(0).astype('Int64')
                logger.info(f"Rounded {col} in Sales sheet to whole numbers")

        return df

    @staticmethod
    def _standardize_price_columns(df: pd.DataFrame) -> pd.DataFrame:
        """Standardize column names for price list."""
        # Convert all column names to strings first, then lowercase
        df.columns = df.columns.astype(str).str.lower().str.strip()
        
        logger.info(f"Price list original columns: {list(df.columns)}")
        
        column_mapping = {
            'master model': 'Master_Model',
            'master_model': 'Master_Model',
            'mastermodel': 'Master_Model',
            'master': 'Master_Model',
            'model': 'Master_Model',
            'valid from': 'Valid_From',
            'valid_from': 'Valid_From',
            'validfrom': 'Valid_From',
            'valid to': 'Valid_To',
            'valid_to': 'Valid_To',
            'validto': 'Valid_To',
            'purchase price': 'Purchase_Price',
            'purchase_price': 'Purchase_Price',
            'purchaseprice': 'Purchase_Price',
            'purchase invoice': 'Purchase_Price',
            'purchase_invoice': 'Purchase_Price',
            'purchaseinvoice': 'Purchase_Price',
            'pre gst price': 'Pre_GST_Price',
            'pre_gst_price': 'Pre_GST_Price',
            'pregstprice': 'Pre_GST_Price',
        }

        df = df.rename(columns=column_mapping)
        
        logger.info(f"Price list columns after mapping: {list(df.columns)}")
        
        # If Purchase_Price is still missing, try broader fuzzy matching
        if 'Purchase_Price' not in df.columns:
            for col in df.columns:
                col_lower = str(col).lower()
                # Look for any column containing 'price' or 'invoice' (but not 'pre gst')
                if ('price' in col_lower or 'invoice' in col_lower) and 'pre' not in col_lower and 'gst' not in col_lower:
                    df = df.rename(columns={col: 'Purchase_Price'})
                    logger.info(f"Mapped '{col}' to 'Purchase_Price' via fuzzy matching")
                    break
        
        # If Master_Model is still missing, try fuzzy matching
        if 'Master_Model' not in df.columns:
            for col in df.columns:
                col_lower = str(col).lower()
                if 'model' in col_lower or 'master' in col_lower:
                    df = df.rename(columns={col: 'Master_Model'})
                    logger.info(f"Mapped '{col}' to 'Master_Model' via fuzzy matching")
                    break
        
        return df

    @staticmethod
    def _standardize_scheme_columns(df: pd.DataFrame) -> pd.DataFrame:
        """Standardize column names for scheme file."""
        # Convert all column names to strings first, then lowercase for consistent matching
        df.columns = df.columns.astype(str).str.lower().str.strip()
        
        logger.info(f"Scheme columns after lowercase: {list(df.columns)}")

        column_mapping = {
            'master model': 'Master_Model',
            'master_model': 'Master_Model',
            'master': 'Master_Model',
            'model': 'Master_Model',
            'start date': 'Scheme_Start_Date',
            'start_date': 'Scheme_Start_Date',
            'scheme start date': 'Scheme_Start_Date',
            'scheme_start_date': 'Scheme_Start_Date',
            'start': 'Scheme_Start_Date',
            'from date': 'Scheme_Start_Date',
            'end date': 'Scheme_End_Date',
            'end_date': 'Scheme_End_Date',
            'scheme end date': 'Scheme_End_Date',
            'scheme_end_date': 'Scheme_End_Date',
            'end': 'Scheme_End_Date',
            'to date': 'Scheme_End_Date',
            'pct scheme -1': 'Pct_Scheme_1',
            'pct scheme-1': 'Pct_Scheme_1',
            'pct scheme 1': 'Pct_Scheme_1',
            'pctscheme-1': 'Pct_Scheme_1',
            'pctscheme1': 'Pct_Scheme_1',
            'pct scheme -1 (a)': 'Pct_Scheme_1_A',
            'pct scheme -1(a)': 'Pct_Scheme_1_A',
            'pct scheme -1 a': 'Pct_Scheme_1_A',
            'pct scheme-1 (a)': 'Pct_Scheme_1_A',
            'pct scheme-1(a)': 'Pct_Scheme_1_A',
            'pct scheme-1 a': 'Pct_Scheme_1_A',
            'pct scheme 1 (a)': 'Pct_Scheme_1_A',
            'pct scheme 1(a)': 'Pct_Scheme_1_A',
            'pct scheme 1 a': 'Pct_Scheme_1_A',
            'pctscheme-1(a)': 'Pct_Scheme_1_A',
            'pctscheme1(a)': 'Pct_Scheme_1_A',
            'pct scheme -1 (b)': 'Pct_Scheme_1_B',
            'pct scheme -1(b)': 'Pct_Scheme_1_B',
            'pct scheme -1 b': 'Pct_Scheme_1_B',
            'pct scheme-1 (b)': 'Pct_Scheme_1_B',
            'pct scheme-1(b)': 'Pct_Scheme_1_B',
            'pct scheme-1 b': 'Pct_Scheme_1_B',
            'pct scheme 1 (b)': 'Pct_Scheme_1_B',
            'pct scheme 1(b)': 'Pct_Scheme_1_B',
            'pct scheme 1 b': 'Pct_Scheme_1_B',
            'pctscheme-1(b)': 'Pct_Scheme_1_B',
            'pctscheme1(b)': 'Pct_Scheme_1_B',
            'pct scheme -1 (c)': 'Pct_Scheme_1_C',
            'pct scheme -1(c)': 'Pct_Scheme_1_C',
            'pct scheme -1 c': 'Pct_Scheme_1_C',
            'pct scheme-1 (c)': 'Pct_Scheme_1_C',
            'pct scheme-1(c)': 'Pct_Scheme_1_C',
            'pct scheme-1 c': 'Pct_Scheme_1_C',
            'pct scheme 1 (c)': 'Pct_Scheme_1_C',
            'pct scheme 1(c)': 'Pct_Scheme_1_C',
            'pct scheme 1 c': 'Pct_Scheme_1_C',
            'pctscheme-1(c)': 'Pct_Scheme_1_C',
            'pctscheme1(c)': 'Pct_Scheme_1_C',
            'pct scheme -2': 'Pct_Scheme_2',
            'pct scheme-2': 'Pct_Scheme_2',
            'pct scheme 2': 'Pct_Scheme_2',
            'pctscheme-2': 'Pct_Scheme_2',
            'pctscheme2': 'Pct_Scheme_2',
            'pct scheme -2 (a)': 'Pct_Scheme_2_A',
            'pct scheme -2(a)': 'Pct_Scheme_2_A',
            'pct scheme -2 a': 'Pct_Scheme_2_A',
            'pct scheme-2 (a)': 'Pct_Scheme_2_A',
            'pct scheme-2(a)': 'Pct_Scheme_2_A',
            'pct scheme-2 a': 'Pct_Scheme_2_A',
            'pct scheme 2 (a)': 'Pct_Scheme_2_A',
            'pct scheme 2(a)': 'Pct_Scheme_2_A',
            'pct scheme 2 a': 'Pct_Scheme_2_A',
            'pctscheme-2(a)': 'Pct_Scheme_2_A',
            'pctscheme2(a)': 'Pct_Scheme_2_A',
            'pct scheme -3': 'Pct_Scheme_3',
            'pct scheme-3': 'Pct_Scheme_3',
            'pct scheme 3': 'Pct_Scheme_3',
            'pctscheme-3': 'Pct_Scheme_3',
            'pctscheme3': 'Pct_Scheme_3',
            'pct scheme -4': 'Pct_Scheme_4',
            'pct scheme-4': 'Pct_Scheme_4',
            'pct scheme 4': 'Pct_Scheme_4',
            'pctscheme-4': 'Pct_Scheme_4',
            'pctscheme4': 'Pct_Scheme_4',
            'pct scheme -5': 'Pct_Scheme_5',
            'pct scheme-5': 'Pct_Scheme_5',
            'pct scheme 5': 'Pct_Scheme_5',
            'pctscheme-5': 'Pct_Scheme_5',
            'pctscheme5': 'Pct_Scheme_5',
            'pct scheme -6': 'Pct_Scheme_6',
            'pct scheme-6': 'Pct_Scheme_6',
            'pct scheme 6': 'Pct_Scheme_6',
            'pctscheme-6': 'Pct_Scheme_6',
            'pctscheme6': 'Pct_Scheme_6',
            'flat schme': 'Flat_Scheme',
            'flat scheme': 'Flat_Scheme',
            'flat_scheme': 'Flat_Scheme',
            'flat': 'Flat_Scheme',
            'flat payout': 'Flat_Scheme',
            'flatschme': 'Flat_Scheme',
            'condition-1': 'Condition_1',
            'condition -1': 'Condition_1',
            'condition_1': 'Condition_1',
            'condition 1': 'Condition_1',
            'condition1': 'Condition_1',
            'condition': 'Condition_1',
            'condition-2': 'Condition_2',
            'condition -2': 'Condition_2',
            'condition_2': 'Condition_2',
            'condition 2': 'Condition_2',
            'condition2': 'Condition_2',
            'flat start date': 'Flat_Payout_Start_Date',
            'flat_start_date': 'Flat_Payout_Start_Date',
            'flatstartdate': 'Flat_Payout_Start_Date',
            'flat start': 'Flat_Payout_Start_Date',
            'flat end date': 'Flat_Payout_End_Date',
            'flat_end_date': 'Flat_Payout_End_Date',
            'flatenddate': 'Flat_Payout_End_Date',
            'flat end': 'Flat_Payout_End_Date',
            'flate end date': 'Flat_Payout_End_Date',  # Handle typo variation
            'flate_end_date': 'Flat_Payout_End_Date',  # Handle typo variation
        }

        df = df.rename(columns=column_mapping)
        
        logger.info(f"Scheme columns after mapping: {list(df.columns)}")
        
        # If still missing required columns, try to find them by pattern matching
        if 'Master_Model' not in df.columns:
            for col in df.columns:
                if 'model' in col.lower() or 'master' in col.lower():
                    df = df.rename(columns={col: 'Master_Model'})
                    logger.info(f"Mapped '{col}' to 'Master_Model'")
                    break
        
        if 'Scheme_Start_Date' not in df.columns:
            for col in df.columns:
                if 'start' in col.lower() or 'from' in col.lower():
                    df = df.rename(columns={col: 'Scheme_Start_Date'})
                    logger.info(f"Mapped '{col}' to 'Scheme_Start_Date'")
                    break
        
        if 'Scheme_End_Date' not in df.columns:
            for col in df.columns:
                if 'end' in col.lower() or 'to' in col.lower():
                    df = df.rename(columns={col: 'Scheme_End_Date'})
                    logger.info(f"Mapped '{col}' to 'Scheme_End_Date'")
                    break

        # Convert percentage columns from whole numbers to decimals (e.g., 2.5 -> 0.025)
        # Store original percentage values for display
        percentage_columns = ['Pct_Scheme_1', 'Pct_Scheme_2', 'Pct_Scheme_3', 'Pct_Scheme_4', 'Pct_Scheme_5', 'Pct_Scheme_6']
        for col in percentage_columns:
            if col in df.columns:
                logger.info(f"Scheme column {col} sample value: {df[col].iloc[0] if len(df) > 0 else 'empty'}")
                # Keep original percentage values, don't convert to decimal
                # The calculation will handle the conversion

        return df

    @staticmethod
    def _standardize_drop_columns(df: pd.DataFrame) -> pd.DataFrame:
        """Standardize column names for drop dump."""
        # Convert all column names to strings first, then lowercase for consistent mapping
        df.columns = df.columns.astype(str).str.lower().str.strip()

        column_mapping = {
            'imei': 'IMEI',
            'drop amount': 'Drop_Amount',
            'drop_amount': 'Drop_Amount',
            'dropamount': 'Drop_Amount',
            'drop': 'Drop_Amount',
            'amount': 'Drop_Amount',
        }

        df = df.rename(columns=column_mapping)

        logger.info(f"Drop dump columns after standardization: {list(df.columns)}")

        # Validate required columns exist
        if 'IMEI' not in df.columns:
            # Try to find IMEI column by partial matching
            for col in df.columns:
                if 'imei' in col.lower():
                    df = df.rename(columns={col: 'IMEI'})
                    logger.info(f"Mapped '{col}' to 'IMEI'")
                    break
            
            if 'IMEI' not in df.columns:
                raise ValueError("Drop dump sheet must contain 'IMEI' or 'imei' column")

        if 'Drop_Amount' not in df.columns:
            # Try to find Drop_Amount column by partial matching
            for col in df.columns:
                if 'drop' in col.lower() or 'amount' in col.lower():
                    df = df.rename(columns={col: 'Drop_Amount'})
                    logger.info(f"Mapped '{col}' to 'Drop_Amount'")
                    break
            
            if 'Drop_Amount' not in df.columns:
                logger.warning("Drop dump sheet missing 'Drop_Amount' column, assuming 0 for all records")
                df['Drop_Amount'] = 0

        return df

    @staticmethod
    def save_results_to_workbook(
        input_file: str,
        processed_data: pd.DataFrame,
        pivot_data: pd.DataFrame,
        output_file: Optional[str] = None
    ) -> str:
        """
        Save processed results as a beautifully formatted Excel file.

        Args:
            input_file: Path to the input workbook
            processed_data: Processed sales data with calculations
            pivot_data: Distributor pivot table data
            output_file: Optional output file path

        Returns:
            Path to the saved file in output folder
        """
        try:
            # Generate output filename
            if output_file is None:
                base_name = os.path.splitext(os.path.basename(input_file))[0]
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file = f"{base_name}_processed_{timestamp}.xlsx"

            # Prepare dataframes for Excel export
            dataframes = {
                'Processed Sales Data': processed_data
            }

            # Add distributor pivot if available
            if pivot_data is not None and not pivot_data.empty:
                dataframes['Distributor Pivot'] = pivot_data

            # Create beautifully formatted Excel file
            filepath = DataLoader.create_beautiful_excel(dataframes, output_file)

            logger.info(f"Excel results saved to {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Error saving results: {str(e)}")
            raise

    @staticmethod
    def _create_summary_sheet(processed_data: pd.DataFrame, pivot_data: pd.DataFrame) -> pd.DataFrame:
        """Create a summary sheet with key metrics."""
        summary = []

        # Overall metrics (with safe column access)
        total_incentives = processed_data.get('Total_Incentive_Received', pd.Series()).sum()
        total_nlc = processed_data.get('NLC', processed_data.get('Calculated_NLC', pd.Series())).sum()
        total_margin = processed_data.get('Margin', processed_data.get('Calculated_Margin', pd.Series())).sum()
        total_records = len(processed_data)

        summary.extend([
            ['SIAT PROCESSING SUMMARY'],
            ['Generated on:', pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')],
            [''],
            ['OVERALL METRICS'],
            ['Total Records Processed', total_records],
            ['Total Incentives Paid', f"₹{total_incentives:,.0f}"],
            ['Total Net Landing Cost', f"₹{total_nlc:,.0f}"],
            ['Total Margin Earned', f"₹{total_margin:,.0f}"],
            ['Average Incentive per Record', f"₹{total_incentives/total_records:,.0f}" if total_records > 0 else 'N/A'],
            [''],
            ['DISTRIBUTOR BREAKDOWN'],
            ['Distributor', 'Total Incentives', 'Margin %'],
        ])

        # Add distributor data
        if pivot_data is not None and not pivot_data.empty:
            for _, row in pivot_data.iterrows():
                summary.append([
                    row.get('Distributor', 'Unknown'),
                    f"₹{row.get('Total_Incentive_Received', 0):,.0f}",
                    f"{row.get('Margin_Percentage', row.get('Margin', 0)):.1f}%"
                ])
        else:
            summary.append(['No distributor data available', '', ''])

        return pd.DataFrame(summary)

    @staticmethod
    def create_beautiful_excel(dataframes: Dict[str, pd.DataFrame], filename: str = None) -> str:
        """
        Create a professionally formatted Excel file with multiple sheets.

        Args:
            dataframes: Dictionary of sheet names to DataFrames
            filename: Output filename (auto-generated if None)

        Returns:
            Path to the created Excel file
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"SIAT_Report_{timestamp}.xlsx"

        # Create output folder if it doesn't exist
        output_folder = "output"
        os.makedirs(output_folder, exist_ok=True)
        filepath = os.path.join(output_folder, filename)

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        for sheet_name, df in dataframes.items():
            ws = wb.create_sheet(title=sheet_name)

            # Write data
            DataLoader._write_formatted_sheet(ws, df, sheet_name)

        # Save workbook
        wb.save(filepath)
        logger.info(f"Beautiful Excel file created: {filepath}")

        return filepath

    @staticmethod
    def _write_formatted_sheet(ws, df: pd.DataFrame, sheet_title: str):
        """Write a DataFrame to a worksheet with professional formatting."""
        # Define styles
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        total_font = Font(name='Calibri', size=11, bold=True, color='000000')
        total_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
        total_alignment = Alignment(horizontal='right', vertical='center')

        data_font = Font(name='Calibri', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Add title row
        title_font = Font(name='Calibri', size=14, bold=True, color='366092')
        title_alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A1:Z1')
        title_cell = ws['A1']
        title_cell.value = f"SIAT Report - {sheet_title}"
        title_cell.font = title_font
        title_cell.alignment = title_alignment

        # Add metadata
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(name='Calibri', size=9, italic=True)
        
        # Calculate totals for numeric columns
        headers = df.columns.tolist()
        totals_row = []
        for col in headers:
            # Check if column is numeric (including Int64, Float64, int64, float64)
            if pd.api.types.is_numeric_dtype(df[col]):
                total_val = df[col].sum()
                totals_row.append(total_val)
            else:
                totals_row.append('')
        
        # Write totals row (row 3)
        totals_row[0] = 'TOTAL'  # First column shows "TOTAL" label
        for col_num, total_val in enumerate(totals_row, 1):
            cell = ws.cell(row=3, column=col_num, value=total_val)
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = total_alignment
            cell.border = border
            if isinstance(total_val, (int, float)):
                cell.number_format = '#,##0.00'

        # Write headers (row 4)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Write data with formatting (starting from row 5)
        for row_num, row in enumerate(df.itertuples(index=False), 5):
            # Alternate row colors
            if row_num % 2 == 0:
                row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            else:
                row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                cell.fill = row_fill

                # Apply data type specific formatting
                column_name = str(headers[col_num-1])
                DataLoader._format_cell_value(cell, value, column_name)

        # Auto-adjust column widths
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = max(
                len(str(header)),
                max([len(str(row[col_num-1])) for row in df.itertuples(index=False)] + [0])
            )
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A5'

        # Add auto-filter
        if len(df) > 0:
            ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}4"

    @staticmethod
    def _format_cell_value(cell, value, column_name: str):
        """Apply appropriate formatting based on data type and column name."""
        from openpyxl.styles import NamedStyle

        if pd.isna(value):
            return

        # IMEI formatting - treat as text, no commas
        if 'imei' in column_name.lower():
            cell.value = str(value).replace(',', '')
            cell.number_format = '@'  # Text format
            return

        # Date formatting
        if 'date' in column_name.lower() or 'Date' in column_name:
            try:
                if isinstance(value, str):
                    # Try to parse string dates
                    cell.value = pd.to_datetime(value).date()
                cell.number_format = 'YYYY-MM-DD'
            except:
                pass

        # Currency formatting
        elif any(keyword in column_name.lower() for keyword in ['price', 'amount', 'incentive', 'margin', 'cost', 'total']):
            try:
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
            except:
                pass

        # Number formatting
        elif isinstance(value, (int, float)):
            if value == int(value):
                cell.number_format = '#,##0'
            else:
                cell.number_format = '#,##0.00'